from pathlib import Path
import csv
import io
import re
import sys

from openpyxl import load_workbook


SOURCE_SHEET = "INVENTARIO VALORIZADO"
MOVEMENT_SHEET = "LMA"


def number(value):
    if value in (None, ""):
        return 0
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return 0


def clean(value):
    return "" if value is None else str(value).strip()


def sql_literal(value):
    return "'" + str(value).replace("'", "''") + "'"


def date_text(value):
    if hasattr(value, "date"):
        return value.date().isoformat()
    return clean(value)


def write_copy_block(handle, table_name, columns, rows):
    handle.write(f"copy {table_name} ({', '.join(columns)}) from stdin with (format csv, header true);\n")
    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow(columns)
    writer.writerows(rows)
    handle.write(buffer.getvalue())
    handle.write("\\.\n\n")


def build_rows(source):
    workbook = load_workbook(source, read_only=True, data_only=True)
    inventory_sheet = workbook[SOURCE_SHEET]
    items = []
    codes = set()

    for index, row in enumerate(inventory_sheet.iter_rows(min_row=2, values_only=True), start=1):
        (
            warehouse,
            ceco,
            status,
            group,
            code,
            description,
            entries,
            exits,
            balance,
            unit_value,
            total_value,
        ) = row[:11]

        if not code or not description:
            continue

        code_text = clean(code)
        item_id = re.sub(r"[^a-zA-Z0-9]+", "-", f"{code_text}-{index}").strip("-").lower()
        codes.add(code_text)
        items.append([
            item_id,
            clean(warehouse),
            clean(ceco),
            clean(status),
            clean(group),
            code_text,
            clean(description),
            number(entries),
            number(exits),
            number(balance),
            number(unit_value),
            round(number(total_value)),
        ])

    movement_sheet = workbook[MOVEMENT_SHEET]
    movements = []

    for row in movement_sheet.iter_rows(min_row=2, values_only=True):
        code = clean(row[0])
        if code not in codes:
            continue

        movements.append([
            code,
            date_text(row[2]),
            clean(row[3]),
            clean(row[4]),
            clean(row[5]),
            number(row[6]),
            number(row[7]),
            number(row[8]),
            number(row[9]) if len(row) > 9 else 0,
        ])

    return items, movements


def main():
    if len(sys.argv) != 3:
        raise SystemExit(
            "Usage: python3 scripts/generate-supabase-seed.py <inventory.xlsx> <output.sql>"
        )

    source = Path(sys.argv[1])
    output = Path(sys.argv[2])
    items, movements = build_rows(source)
    output.parent.mkdir(parents=True, exist_ok=True)

    inventory_columns = [
        "id",
        "warehouse",
        "ceco",
        "status",
        "item_group",
        "code",
        "description",
        "entries",
        "exits",
        "balance",
        "unit_value",
        "total_value",
    ]
    movement_columns = [
        "item_code",
        "movement_date",
        "document",
        "document_number",
        "warehouse",
        "entry_qty",
        "exit_qty",
        "balance_qty",
        "unit_value",
    ]

    with output.open("w", encoding="utf-8", newline="") as handle:
        handle.write("-- Generated from the real ToolTrack Excel inventory.\n")
        handle.write("-- Run after supabase/schema.sql. Review before executing in production.\n")
        handle.write("begin;\n\n")
        handle.write("create temp table tooltrack_inventory_stage (like public.inventory_items including defaults);\n")
        handle.write("create temp table tooltrack_movement_stage (like public.inventory_movements including defaults);\n\n")
        write_copy_block(handle, "tooltrack_inventory_stage", inventory_columns, items)
        write_copy_block(handle, "tooltrack_movement_stage", movement_columns, movements)
        handle.write(
            "insert into public.inventory_imports "
            "(source_file, row_count, movement_row_count, status) values "
            f"({sql_literal(source.name)}, {len(items)}, {len(movements)}, 'completed');\n\n"
        )
        handle.write(
            "insert into public.inventory_items "
            f"({', '.join(inventory_columns)}) "
            f"select {', '.join(inventory_columns)} from tooltrack_inventory_stage "
            "on conflict (id) do update set "
            "warehouse = excluded.warehouse, "
            "ceco = excluded.ceco, "
            "status = excluded.status, "
            "item_group = excluded.item_group, "
            "code = excluded.code, "
            "description = excluded.description, "
            "entries = excluded.entries, "
            "exits = excluded.exits, "
            "balance = excluded.balance, "
            "unit_value = excluded.unit_value, "
            "total_value = excluded.total_value, "
            "updated_at = now();\n\n"
        )
        handle.write("truncate table public.inventory_movements restart identity;\n")
        handle.write(
            "insert into public.inventory_movements "
            f"({', '.join(movement_columns)}) "
            f"select {', '.join(movement_columns)} from tooltrack_movement_stage;\n\n"
        )
        handle.write("commit;\n")

    print(f"Wrote {len(items)} inventory rows and {len(movements)} movements to {output}")


if __name__ == "__main__":
    main()
