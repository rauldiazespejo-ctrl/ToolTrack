from pathlib import Path
import json
import re
import sys

from openpyxl import load_workbook


SOURCE_SHEET = "INVENTARIO VALORIZADO"
MOVEMENT_SHEET = "LMA"


def number(value):
    if value in (None, ""):
        return 0
    try:
        return round(float(value), 4)
    except (TypeError, ValueError):
        return 0


def clean(value):
    return "" if value is None else str(value).strip()


def main():
    source = Path(sys.argv[1])
    output = Path(sys.argv[2])
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet = workbook[SOURCE_SHEET]
    items = []

    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        (
            warehouse,
            ceco,
            status,
            group,
            code,
            description,
            entries,
            exits,
            balance,
            unit_value,
            total_value,
        ) = row[:11]
        if not code or not description:
            continue

        item_id = re.sub(r"[^a-zA-Z0-9]+", "-", f"{code}-{index}").strip("-").lower()
        items.append({
            "id": item_id,
            "warehouse": clean(warehouse),
            "ceco": ceco,
            "status": clean(status),
            "group": clean(group),
            "code": clean(code),
            "description": clean(description),
            "entries": number(entries),
            "exits": number(exits),
            "balance": number(balance),
            "unitValue": number(unit_value),
            "totalValue": round(number(total_value)),
        })

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(items, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {len(items)} inventory rows to {output}")

    movement_output = output.with_name("movements.generated.json")
    codes = {item["code"] for item in items}
    movement_sheet = workbook[MOVEMENT_SHEET]
    movements = {code: {"movementCount": 0, "lastMovements": []} for code in codes}

    for row in movement_sheet.iter_rows(min_row=2, values_only=True):
        code = clean(row[0])
        if code not in movements:
            continue

        date = row[2]
        if hasattr(date, "isoformat"):
            date_text = date.date().isoformat()
            sort_key = date.isoformat()
        else:
            date_text = clean(date)
            sort_key = date_text

        event = {
            "date": date_text,
            "document": clean(row[3]),
            "number": clean(row[4]),
            "warehouse": clean(row[5]),
            "entry": number(row[6]),
            "exit": number(row[7]),
            "balance": number(row[8]),
            "_sort": sort_key,
        }

        bucket = movements[code]
        bucket["movementCount"] += 1
        bucket["lastMovements"].append(event)
        bucket["lastMovements"].sort(key=lambda item: item["_sort"], reverse=True)
        bucket["lastMovements"] = bucket["lastMovements"][:3]

    for bucket in movements.values():
        for event in bucket["lastMovements"]:
            event.pop("_sort", None)

    movement_output.write_text(
        json.dumps(movements, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"Wrote movement summaries to {movement_output}")


if __name__ == "__main__":
    main()
